import requests
import csv
import pandas as pd
import openpyxl
import re
import os
import time 

from bs4 import BeautifulSoup as bs
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

#To define the driver as headless
op = webdriver.ChromeOptions()
op.add_argument('headless')

#Function to write in the spreadsheet
def writeDF(df, wb, f):
  ws = wb['Sheet']

  #Creating an array of strings needed filter out of the report
  check_list = [
    'Required Courses',
    'YSU 1500', 'or SS 1500',
    'or HONR 1500',
    'FIRST YEAR REQUIREMENT -STUDENT SUCCESS SEMINAR',
    'FIRST YEAR REQUIREMENT -STUDENT SUCCESS',
    'OR'
  ]

  row_index = len(ws['C'])
  for x in range(len(df)):
    if df.iloc[x,0] not in check_list:
      value = df.iloc[x,0]
      if(ws.cell(column = 3, row = row_index)).value is None:
        ws.cell(column = 3, row = row_index, value = value)
      else:
        ws.cell(column = 3, row = row_index + 1, value = value)
      wb.save(filename = f)
      row_index += 1
      #newRowLocation = ws.max_row + 1

#A function to pull & store the links of each programs into an array
def url(link_list, year):
  if(year != curr_year):
    url = 'https://catalog.ysu.edu/archives/' + year

  else:
    url = 'https://catalog.ysu.edu/'

  driver = webdriver.Chrome(options=op)
  driver.get(url)

  #Click the A-Z Programs
  if(year != curr_year):
    driver.find_element_by_xpath("//*[@href=" +"'"+ '/archives/' + year +"/programs/" +"'" +"]").click()
  else:
    driver.find_element_by_xpath("//*[@href='/programs/']").click()

  #Getting the page source of the url driver is on
  r = driver.page_source.encode('utf-8').strip()

  #Using BeautifulSoup to pull the table to access all the programs link
  soup = bs(r, 'html.parser')
  programs_table = soup.find('table', {"class":"sc_sctable tbl_degrees sorttable"})

  #Creating a list to avoid the pre-biology and pre-chemistry links which are invalid and makes program slow
  skip_list = [
    'https://catalog.ysu.edu/archives/'+year+'/undergraduate/colleges-programs/college-science-technology-engineering-mathematics/department-chemistry/',
    'https://catalog.ysu.edu/archives/'+year+'/undergraduate/colleges-programs/college-science-technology-engineering-mathematics/department-biological-sciences/',
    'https://catalog.ysu.edu/undergraduate/colleges-programs/college-science-technology-engineering-mathematics/department-biological-sciences/',
    'https://catalog.ysu.edu/undergraduate/colleges-programs/college-science-technology-engineering-mathematics/department-chemistry/',
  ]

  #Iterating through the programs_table to find all 'a' to pull the href for the programs and adding it to programs_link array
  for a in programs_table.find_all('a'):
    if ('https://catalog.ysu.edu'+a['href']) in skip_list:
      continue

    link_list.append('https://catalog.ysu.edu'+a['href'])

  driver.quit()

def containerIdentifier(driver):
  #Using find_element's' to make the variable an array to be able to compare in if else for the content and return proper identifier
  curriculumSheet = driver.find_elements_by_id('curriculumsheettexttab')
  certificateReq = driver.find_elements_by_id('certificaterequirementstexttab')
  degreeReq = driver.find_elements_by_id('newitemtexttab')
  fourYearPlan = driver.find_elements_by_id('fouryearplantexttab')
  twoYearPlan = driver.find_elements_by_id('twoyearplantexttab')
  degreeReq2 = driver.find_elements_by_id('degreerequirementstexttab')

  if (curriculumSheet):
    return 'curriculumsheettexttab'

  elif (degreeReq):
    return 'newitemtexttab'

  elif (degreeReq2):
    return 'degreerequirementstexttab'

  elif (certificateReq):
    return 'certificaterequirementstexttab'

  elif (fourYearPlan):
    return 'fouryearplantexttab'

  elif (twoYearPlan):
    return 'twoyearplantexttab'
    
  else:
    return 'none'

#This function compares the two links using pandas dataframe
def compare(html1, html2):
  try:
    #reading the html as pandas dataframe
    df1 = pd.read_html(html1)
  except ValueError:
    print("\n '" + Title1 + "' does not have a table")
    return
  
  #The dataframe for this program is in df[1] and everthing else is in df[0]
  if(Title1 == "Master of Arts in Art Education"):
    df1 = df1[1]
  else:
    df1 = df1[0]

  try:
    df2 = pd.read_html(html2)

  except ValueError:
    print("\n '" + Title2 + "' does not have a table")
    return
  
  #The dataframe for this program is in df[1] and everthing else is in df[0], same as Title1
  if(Title2 == "Master of Arts in Art Education"):
    df2 = df2[1]
  else:
    df2 = df2[0]

  #if the pulled dataframe does not have a title in proper order these lines adds them
  if (list(df1.columns.values) != ['COURSE', 'TITLE', 'S.H.']):
    df1.columns = ['COURSE', 'TITLE', 'S.H.']
  
  if (list(df2.columns.values) != ['COURSE', 'TITLE', 'S.H.']):
    df2.columns = ['COURSE','TITLE','S.H.']

  #COMPARISON LINE
  #This line compares two dataframes COURSE column and checks if one is in the other
  df1['compare'] = df1['COURSE'].isin(df2['COURSE'])
  df2['compare'] = df2['COURSE'].isin(df1['COURSE'])

  #If there is even one false in compare column of df1 filter the ones with false, else set the dataframe null or empty
  if False in df1['compare']:
    df1 = df1[df1['compare'] == False]
  else: 
    df1 = pd.DataFrame()
  
  #If there is even one false in compare column of df1 filter the ones with false, else set the dataframe null or empty
  if False in df2['compare']:
    df2 = df2[df2['compare'] == False]
  else: 
    df2 = pd.DataFrame()

  #If both df1 and df2 are empty get out of the function since we dont need to do any comparisons
  if(df1.empty) and (df2.empty):
    return

  # Drop the NaN in the COURSE column to avoid adding blank spaces to the spreadsheet
  df1 = df1.dropna(subset = ['COURSE'])
  df2 = df2.dropna(subset = ['COURSE'])

  #Create new workbook
  wb = Workbook()
  ws = wb.active
  f = 'Changes in program.xlsx'
  if not (os.path.isfile(f)):
    Workbook().save(f)

  wb = openpyxl.load_workbook(filename = f)
  ws = wb['Sheet']
  ws.cell(column = 1, row = 1, value = 'TITLE')
  ws.cell(column = 2, row = 1, value = 'STATUS')
  ws.cell(column = 3, row = 1, value = 'CHANGES')

  titleLocation = ws.max_row + 1
  ws.cell(column = 1, row = titleLocation, value = Title1)
  ws.cell(column = 2, row = titleLocation, value = 'Changed')

  writeDF(df1, wb, f)
  writeDF(df2, wb, f)

  wb.save(filename = f)

#Ask the user to input the current year which is used in comparison later
curr_year = input("\nEnter current year (format: 20XX-20XX): ")

#Ask the user to input the years they want to compare separated with an &, & is later used to split the two years
comparison = input("\nEnter the year's you want to compare separated by an & (YEAR1 & YEAR2): ")

#curr_year = "2020-2021"
#comparison = "2019-2020&2020-2021"
#comparison = "2020-2021&2019-2020"

year1, year2 = comparison.split('&', 1)
year1 = year1.strip()
year2 = year2.strip()

programs_link_1 = []
programs_link_2 = []

url(programs_link_1, year1)
url(programs_link_2, year2)

#Bool Variable to check if the two programs title are equal
TitleEqual = False

#driver1 = webdriver.Chrome(options=op)
driver1 = webdriver.Chrome()
driver2 = webdriver.Chrome()

#Creating an array to skip the programs that give error or are only present in a particular year which makes the comparison impossible 
skip_programs = [
  'Bachelor of Science in Education in Art Education (PK to 12) Multi-Age License',
  'Biological Sciences',
  'Chemical Sciences',
  'Department of Biological Sciences',
  'Department of Chemistry',
  'Certificate in Data Analytics',
  'Certificate in Data Analytics',
  'Certificate in Homeland Security',
  'Bachelor of Science in Education in Primary/Primary Intervention Specialist Education (P-5)',
  'Department of Art',
  'Minor in Coaching Education P-16',
  'Bachelor of Arts in Dance Management',
  'Associate of Applied Science in Dietetic Technician',
  'Bachelor of Science in Education in Early Childhood Education/Early Childhood Intervention Specialist',
  'Bachelor of Science in Applied Science in Family and Consumer Studies, Consumer Studies Track',
  'Bachelor of Science in Applied Science in Family and Consumer Studies, Family Studies Track',
  'Family and Consumer Studies, Instructor Track'
  'Bachelor of Science in Applied Science in Family and Consumer Studies, Instructor Track',
  'Bachelor of Science in Education in Health Education (PK-12) - Multi-Age License',
  'Bachelor of Science in Business Administration in Information and Supply Chain Management',
  'Associate of Applied Science in Medical Assisting Technology',
  'Bachelor of Science in Education in Physical Education (PK-12) - Multi-Age License'
]

for x in range(len(programs_link_1)):
  #To check if the link is working
  response = requests.get(programs_link_1[x])
  if response.status_code != 200:
    print('\n' + 'This link is bad:' + programs_link_1[x])
    continue

  #if the link is working get the link
  driver1.get(programs_link_1[x])

  Title1 = (driver1.find_element_by_class_name('page-title')).text.strip()

  if(Title1 in skip_programs):
    continue

  clean_title1 = re.sub('\W+',' ', Title1)

  for y in range(len(programs_link_2)):
    #To check if the link is working
    response = requests.get(programs_link_2[y])
    if response.status_code != 200:
      print('\n' + 'This link is bad:' + programs_link_2[y])
      continue

    driver2.get(programs_link_2[y])
    Title2 = driver2.find_element_by_class_name('page-title').text.strip()

    if(Title2 in skip_programs):
      # time.sleep(5)
      programs_link_2.pop(y)
      break

    clean_title2 = re.sub('\W+',' ', Title2)

    if(clean_title1 == clean_title2):
        TitleEqual = True
        #del programs_link_2[y]
        
        programs_link_2.pop(y)
        # programs_link_1.pop(x)
        break

  if(TitleEqual):
    if (containerIdentifier(driver1) != 'none'):
      element1 = driver1.find_element_by_id(containerIdentifier(driver1))
      driver1.execute_script("arguments[0].scrollIntoView();", element1)
      element1.click()

    html1 = driver1.page_source

    if(containerIdentifier(driver2) != 'none'):
      element2 = driver2.find_element_by_id(containerIdentifier(driver2))
      driver2.execute_script("arguments[0].scrollIntoView();", element2)
      element2.click()

    html2 = driver2.page_source
    compare(html1, html2)
  else: 
    print('\n' + Title1 + ' in not present in', year2)